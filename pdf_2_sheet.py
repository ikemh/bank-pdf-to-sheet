import pdfplumber
import pandas as pd
import re
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def extrair_movimentacoes_pdf(caminho_pdf):
    registros = []
    saldo_anterior = None
    with pdfplumber.open(caminho_pdf) as pdf:
        for pagina in pdf.pages:
            texto = pagina.extract_text()
            if not texto:
                continue
            linhas = texto.split('\n')
            for linha in linhas:
                # Procura por SALDO ANTERIOR
                if 'SALDO ANTERIOR' in linha.upper():
                    match_saldo = re.search(r'(-?\d{1,3}(?:\.\d{3})*,\d{2})$', linha)
                    if match_saldo:
                        val_str = match_saldo.group(1).replace('.', '').replace(',', '.')
                        saldo_anterior = float(val_str)
                
                # Filtra somente linhas que iniciem com data
                if not re.match(r'\d{2}/\d{2}/\d{4}', linha):
                    continue
                data = linha[:10]
                restante = linha[11:]
                match = re.search(r'(-?\d{1,3}(?:\.\d{3})*,\d{2})\s+(-?\d{1,3}(?:\.\d{3})*,\d{2})$', restante)
                if match:
                    valor = match.group(1)
                    saldo = match.group(2)
                    descricao = restante[:match.start()].strip()
                    registros.append([data, descricao, valor, saldo])
    df = pd.DataFrame(registros, columns=["Data", "Descrição", "Valor (R$)", "Saldo (R$)"])
    if not df.empty:
        # Converte formato brasileiro para float
        df["Valor (R$)"] = df["Valor (R$)"].str.replace(".", "", regex=False).str.replace(",", ".", regex=False).astype(float)
        df["Saldo (R$)"] = df["Saldo (R$)"].str.replace(".", "", regex=False).str.replace(",", ".", regex=False).astype(float)
    return df, saldo_anterior

def salvar_excel_formatado(df, caminho_saida, saldo_anterior=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Extrato"

    # Escrita da tabela principal (A1 em diante)
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Adiciona a coluna E (Categoria)
    ws["E1"] = "Categoria"
    
    main_table_end = ws.max_row
    
    # Preenche fórmulas da Categoria
    for row_idx in range(2, main_table_end + 1):
        ws[f"E{row_idx}"] = f'=IF(UPPER(TRIM(B{row_idx}))="RESG.APLIC.FIN.AVISO PREV CAPTACAO","Resgate",IF(C{row_idx}>0,"Entrada",IF(C{row_idx}<0,"Saída","")))'

    main_table_cols = 5
    
    # Fonts e Alinhamentos
    header_font_white = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    money_format = u'R$ #,##0.00'
    
    # Bordas
    thin_side = Side(style='thin', color='BFBFBF')
    thick_side = Side(style='medium', color='595959')
    
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_border = Border(left=thin_side, right=thin_side, top=thick_side, bottom=thick_side)
    
    # Cores de Preenchimento (Elegantes)
    main_header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Azul Escuro Profundo
    main_body_fill_alt1 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Branco
    main_body_fill_alt2 = PatternFill(start_color="EDF2F9", end_color="EDF2F9", fill_type="solid") # Azul Muito Claro
    
    col_entrada_header = PatternFill(start_color="375623", end_color="375623", fill_type="solid") # Verde Escuro
    col_entrada_body = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Verde Pastel
    
    col_saida_header = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")   # Vermelho Escuro
    col_saida_body = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # Vermelho Pastel
    
    col_resgate_header = PatternFill(start_color="843C0C", end_color="843C0C", fill_type="solid") # Laranja Escuro / Marrom
    col_resgate_body = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Amarelo Pastel
    
    conf_header_fill = PatternFill(start_color="44546A", end_color="44546A", fill_type="solid")   # Grafite
    conf_body_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")     # Cinza Claro

    # Aplica estilos na tabela principal (com Zebrado)
    for row in ws.iter_rows(min_row=1, max_row=main_table_end, max_col=main_table_cols):
        is_even = row[0].row % 2 == 0
        row_fill = main_body_fill_alt1 if is_even else main_body_fill_alt2
        
        for cell in row:
            if cell.column_letter == 'B':
                cell.alignment = left_align
            else:
                cell.alignment = center_align
                
            if cell.row == 1:
                cell.font = header_font_white
                cell.fill = main_header_fill
                cell.border = header_border
            else:
                cell.border = thin_border
                cell.fill = row_fill
                if cell.column_letter in ['C', 'D']:
                    cell.number_format = money_format

    # Formatação Condicional VIVA na coluna C (Valor)
    from openpyxl.formatting.rule import FormulaRule, CellIsRule
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    
    ws.conditional_formatting.add(f'C2:C{main_table_end}', FormulaRule(formula=['$E2="Saída"'], stopIfTrue=True, fill=red_fill, font=red_font))
    ws.conditional_formatting.add(f'C2:C{main_table_end}', FormulaRule(formula=['$E2="Entrada"'], stopIfTrue=True, fill=green_fill, font=green_font))

    # Congelar cabeçalho e AutoFilter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:E{main_table_end}"

    # Posição inicial para colunas extras
    extra_start_col = main_table_cols + 2
    col_entrada_letter = get_column_letter(extra_start_col)
    col_saida_letter   = get_column_letter(extra_start_col + 1)
    col_resgate_letter = get_column_letter(extra_start_col + 2)
    
    headers_config = [
        (col_entrada_letter, "Entrada (R$)", col_entrada_header, col_entrada_body),
        (col_saida_letter, "Saída (R$)", col_saida_header, col_saida_body),
        (col_resgate_letter, "Resgates (R$)", col_resgate_header, col_resgate_body)
    ]
    
    # Cabeçalhos extras e Estilos
    for col_letter, title, h_fill, b_fill in headers_config:
        ws[f"{col_letter}1"] = title
        ws[f"{col_letter}1"].font = header_font_white
        ws[f"{col_letter}1"].alignment = center_align
        ws[f"{col_letter}1"].fill = h_fill
        ws[f"{col_letter}1"].border = header_border

    # Fórmulas de Entrada, Saída e Resgate
    for row_idx in range(2, main_table_end + 1):
        ws[f"{col_entrada_letter}{row_idx}"] = f'=IF($E{row_idx}="Entrada",$C{row_idx},"")'
        ws[f"{col_saida_letter}{row_idx}"] = f'=IF($E{row_idx}="Saída",ABS($C{row_idx}),"")'
        ws[f"{col_resgate_letter}{row_idx}"] = f'=IF($E{row_idx}="Resgate",$C{row_idx},"")'
        
        for col_letter, _, _, b_fill in headers_config:
            cell = ws[f"{col_letter}{row_idx}"]
            cell.number_format = money_format
            cell.alignment = center_align
            cell.fill = b_fill
            cell.border = thin_border

    # Totais (SUM) elegantes
    total_row = main_table_end + 2
    for col_letter, _, h_fill, _ in headers_config:
        ws[f"{col_letter}{total_row - 1}"] = "Total"
        ws[f"{col_letter}{total_row - 1}"].font = header_font_white
        ws[f"{col_letter}{total_row - 1}"].alignment = center_align
        ws[f"{col_letter}{total_row - 1}"].fill = h_fill
        ws[f"{col_letter}{total_row - 1}"].border = header_border

        ws[f"{col_letter}{total_row}"] = f"=SUM({col_letter}2:{col_letter}{main_table_end})"
        ws[f"{col_letter}{total_row}"].number_format = money_format
        ws[f"{col_letter}{total_row}"].font = bold_font
        ws[f"{col_letter}{total_row}"].alignment = center_align
        ws[f"{col_letter}{total_row}"].fill = conf_body_fill
        ws[f"{col_letter}{total_row}"].border = thin_border

    # Seção de Conferência do Balancete
    conf_start_col = extra_start_col + 4
    col_label_letter = get_column_letter(conf_start_col)
    col_val_letter = get_column_letter(conf_start_col + 1)
    
    conf_row = 2
    # Título da Conferência
    ws[f"{col_label_letter}{conf_row}"] = "Conferência do Balancete"
    ws[f"{col_label_letter}{conf_row}"].font = header_font_white
    ws[f"{col_label_letter}{conf_row}"].fill = conf_header_fill
    ws[f"{col_label_letter}{conf_row}"].alignment = center_align
    ws[f"{col_val_letter}{conf_row}"].fill = conf_header_fill
    ws.merge_cells(f"{col_label_letter}{conf_row}:{col_val_letter}{conf_row}")
    conf_row += 1

    campos = [
        "Saldo anterior",
        "Saldo final",
        "Variação do saldo",
        "Total entradas",
        "Total saídas",
        "Total resgates",
        "Resultado líquido",
        "Diferença de conferência",
        "Status"
    ]
    
    # Prepara as células da conferência
    conf_cells = {}
    for campo in campos:
        ws[f"{col_label_letter}{conf_row}"] = campo
        ws[f"{col_label_letter}{conf_row}"].font = bold_font
        ws[f"{col_label_letter}{conf_row}"].fill = conf_body_fill
        ws[f"{col_label_letter}{conf_row}"].border = thin_border
        
        ws[f"{col_val_letter}{conf_row}"].fill = main_body_fill_alt1
        ws[f"{col_val_letter}{conf_row}"].border = thin_border
        
        conf_cells[campo] = f"{col_val_letter}{conf_row}"
        conf_row += 1
        
    # Fórmulas da Conferência
    if saldo_anterior is not None:
        ws[conf_cells["Saldo anterior"]] = saldo_anterior
    
    ws[conf_cells["Saldo final"]] = f"=D{main_table_end}"
    ws[conf_cells["Variação do saldo"]] = f"={conf_cells['Saldo final']}-{conf_cells['Saldo anterior']}"
    ws[conf_cells["Total entradas"]] = f"={col_entrada_letter}{total_row}"
    ws[conf_cells["Total saídas"]] = f"={col_saida_letter}{total_row}"
    ws[conf_cells["Total resgates"]] = f"={col_resgate_letter}{total_row}"
    ws[conf_cells["Resultado líquido"]] = f"={conf_cells['Total entradas']}+{conf_cells['Total resgates']}-{conf_cells['Total saídas']}"
    ws[conf_cells["Diferença de conferência"]] = f"={conf_cells['Resultado líquido']}-{conf_cells['Variação do saldo']}"
    ws[conf_cells["Status"]] = f'=IF(ABS({conf_cells["Diferença de conferência"]})<0.01,"OK","DIVERGENTE")'

    # Formatação das células de conferência
    for campo, cel in conf_cells.items():
        cell = ws[cel]
        cell.alignment = right_align
        if campo != "Status":
            cell.number_format = money_format
        else:
            cell.font = bold_font
            cell.alignment = center_align

    # Bordas grossas envolta do balancete
    for r_idx in range(2, conf_row):
        ws[f"{col_label_letter}{r_idx}"].border = Border(left=thick_side, right=thin_side, top=ws[f"{col_label_letter}{r_idx}"].border.top, bottom=ws[f"{col_label_letter}{r_idx}"].border.bottom)
        ws[f"{col_val_letter}{r_idx}"].border = Border(left=thin_side, right=thick_side, top=ws[f"{col_val_letter}{r_idx}"].border.top, bottom=ws[f"{col_val_letter}{r_idx}"].border.bottom)
    
    ws[f"{col_label_letter}2"].border = Border(left=thick_side, right=thin_side, top=thick_side, bottom=thin_side)
    ws[f"{col_val_letter}2"].border = Border(left=thin_side, right=thick_side, top=thick_side, bottom=thin_side)
    ws[f"{col_label_letter}{conf_row-1}"].border = Border(left=thick_side, right=thin_side, top=thin_side, bottom=thick_side)
    ws[f"{col_val_letter}{conf_row-1}"].border = Border(left=thin_side, right=thick_side, top=thin_side, bottom=thick_side)

    # Formatação Condicional para o Status
    ws.conditional_formatting.add(conf_cells["Status"], CellIsRule(operator='equal', formula=['"OK"'], stopIfTrue=True, fill=green_fill, font=green_font))
    ws.conditional_formatting.add(conf_cells["Status"], CellIsRule(operator='equal', formula=['"DIVERGENTE"'], stopIfTrue=True, fill=red_fill, font=red_font))

    # Ajuste automático inteligente da largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value is not None:
                val_str = str(cell.value)
                if val_str.startswith('='):
                    val_str = "R$ 99.999,99"
                max_length = max(max_length, len(val_str))
        adjusted_width = max_length + 4
        if column == 'B' and adjusted_width > 45: 
            adjusted_width = 45 # Limite para descrição
        ws.column_dimensions[column].width = adjusted_width

    wb.save(caminho_saida)

def selecionar_pdf():
    caminho = filedialog.askopenfilename(filetypes=[("PDF files", "*.pdf")])
    entrada_var.set(caminho)

def selecionar_saida():
    caminho = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    saida_var.set(caminho)

def converter():
    entrada = entrada_var.get()
    saida = saida_var.get()
    if not entrada or not saida:
        messagebox.showerror("Erro", "Por favor, selecione o PDF de entrada e o caminho de saída.")
        return
    try:
        df, saldo_anterior = extrair_movimentacoes_pdf(entrada)
        salvar_excel_formatado(df, saida, saldo_anterior)
        messagebox.showinfo("Sucesso", "Arquivo convertido e salvo com sucesso!")
    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")

if __name__ == "__main__":
    janela = tk.Tk()
    janela.title("Conversor de Extrato Bancário (PDF → Excel)")

    entrada_var = tk.StringVar()
    saida_var = tk.StringVar()

    tk.Label(janela, text="Selecionar PDF:").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(janela, textvariable=entrada_var, width=50).grid(row=0, column=1, padx=5)
    tk.Button(janela, text="Buscar", command=selecionar_pdf).grid(row=0, column=2, padx=5)

    tk.Label(janela, text="Salvar como:").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    tk.Entry(janela, textvariable=saida_var, width=50).grid(row=1, column=1, padx=5)
    tk.Button(janela, text="Escolher", command=selecionar_saida).grid(row=1, column=2, padx=5)

    tk.Button(janela, text="Converter", command=converter, bg="#4CAF50", fg="white", height=2).grid(row=2, column=1, pady=20)

    janela.mainloop()
